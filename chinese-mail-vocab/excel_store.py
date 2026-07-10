import os
from datetime import date

from openpyxl import Workbook, load_workbook

HEADERS = ['병음', '한자', '한국어 뜻', '누적 횟수', '최근 확인일']

DEFAULT_PATH = os.path.join(
    os.path.expanduser('~'), 'Documents', 'ChineseWordBook', 'chinese_words.xlsx'
)


def ensure_parent_dir(path):
    os.makedirs(os.path.dirname(path), exist_ok=True)


def load_words(path):
    """Returns dict: {한자: {'pinyin':..., 'meaning':..., 'count':int, 'last_seen':str}}"""
    if not os.path.exists(path):
        return {}

    wb = load_workbook(path)
    ws = wb.active
    words = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[1]:
            continue
        pinyin, hanzi, meaning, count, last_seen = (list(row) + [None] * 5)[:5]
        words[hanzi] = {
            'pinyin': pinyin or '',
            'meaning': meaning or '',
            'count': int(count) if count else 0,
            'last_seen': last_seen or '',
        }
    return words


def _is_hangul_meaning(s):
    return any('가' <= c <= '힣' for c in str(s))


def merge_words(existing, new_entries):
    """new_entries: list of (hanzi, pinyin, meaning). Returns updated dict (mutates a copy)."""
    merged = dict(existing)
    today = date.today().isoformat()

    for hanzi, pinyin, meaning in new_entries:
        if hanzi in merged:
            entry = merged[hanzi]
            entry['count'] += 1
            entry['last_seen'] = today
            # 뜻이 비어있거나 한글이 아닌 값(영어/한자 등)이면 새 뜻으로 교체
            if meaning and not _is_hangul_meaning(entry['meaning']):
                entry['meaning'] = meaning
            if not entry['pinyin'] and pinyin:
                entry['pinyin'] = pinyin
        else:
            merged[hanzi] = {
                'pinyin': pinyin,
                'meaning': meaning,
                'count': 1,
                'last_seen': today,
            }
    return merged


def write_words(path, words):
    """words: dict {한자: {...}}. Sorted by pinyin ascending (tone-mark-aware code point order)."""
    ensure_parent_dir(path)

    wb = Workbook()
    ws = wb.active
    ws.title = '단어장'
    ws.append(HEADERS)

    ordered = sorted(words.items(), key=lambda item: item[1]['pinyin'])
    for hanzi, entry in ordered:
        ws.append([entry['pinyin'], hanzi, entry['meaning'], entry['count'], entry['last_seen']])

    try:
        wb.save(path)
    except PermissionError:
        raise PermissionError(
            '엑셀 파일이 열려 있어 저장할 수 없습니다. 파일을 닫고 다시 시도해주세요.'
        )


def add_words_and_save(new_entries, path=DEFAULT_PATH):
    """new_entries: list of (hanzi, pinyin, meaning). Loads, merges, sorts, saves. Returns merged dict."""
    existing = load_words(path)
    merged = merge_words(existing, new_entries)
    write_words(path, merged)
    return merged
